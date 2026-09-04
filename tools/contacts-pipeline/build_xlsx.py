"""Build the deliverable workbook: original tabs + Contacts + Companies + README."""
import csv
import json
import os
import shutil
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "/tmp/claude-0/-home-user-repflow/aa11d0cc-5220-5a93-8866-4882d0bd9d23/scratchpad"
OUT = f"{BASE}/repflow_contacts_master.xlsx"

contacts = list(csv.DictReader(open(f"{BASE}/final_contacts.csv")))
companies = list(csv.DictReader(open(f"{BASE}/final_companies.csv")))

wb = openpyxl.load_workbook(f"{BASE}/hvac_reps.xlsx")  # original workbook (all tabs)
for name in ["Contacts", "Companies (enriched)", "README"]:
    if name in wb.sheetnames:
        del wb[name]

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_table(ws, rows, cols, widths=None):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    for i, c in enumerate(cols, 1):
        w = (widths or {}).get(c)
        if not w:
            sample = [len(str(r.get(c, ""))) for r in rows[:400]] + [len(c)]
            w = min(max(sample) + 2, 45)
        ws.column_dimensions[get_column_letter(i)].width = w


ccols = ["Company", "First Name", "Last Name", "Job Title", "Company URL", "Email", "Phone Number", "LinkedIn",
         "Email Status", "Alt Email", "Phone Type", "Priority", "City", "State", "Company City", "Company State",
         "Vertical", "Targeting Tier", "Seniority", "Department", "Data Sources", "Company ID"]
ws = wb.create_sheet("Contacts", 0)
write_table(ws, contacts, ccols, {"Company": 34, "Job Title": 34, "Email": 34, "LinkedIn": 48, "Company URL": 32})

kcols = list(companies[0].keys())
ws2 = wb.create_sheet("Companies (enriched)", 1)
write_table(ws2, companies, kcols, {"Company": 34, "Notes": 60, "Team Page URLs": 40})

ws3 = wb.create_sheet("README", 2)
n_ver = sum(1 for r in contacts if r["Email Status"].startswith("Verified"))
n_inf = sum(1 for r in contacts if r["Email Status"].startswith("Inferred"))
n_guess = sum(1 for r in contacts if r["Email Status"].startswith("Best guess"))
n_listed = sum(1 for r in contacts if r["Email Status"].startswith("Listed"))
n_direct = sum(1 for r in contacts if r["Phone Type"].startswith("Direct"))
n_li = sum(1 for r in contacts if r["LinkedIn"])
n_comp_with = len({r["Company ID"] for r in contacts})
lines = [
    ["RepFlow rep-agency contact database", ""],
    ["Built", date.today().isoformat()],
    ["", ""],
    ["Contacts tab", f"{len(contacts)} people across {n_comp_with} rep agencies (from {len(companies)} companies in the Master tab)."],
    ["Companies (enriched) tab", "One row per company: resolved website/domain, LinkedIn page, email pattern, headcount signal, contact count, notes."],
    ["", ""],
    ["Email Status meanings", ""],
    ["Verified (Seamless nn%)", f"{n_ver} rows. Email returned by Seamless.AI contact research with its confidence score. Alt Email holds Seamless's secondary candidates."],
    ["Listed (directory/website)", f"{n_listed} rows. Email printed on a public directory listing or the company website."],
    ["Inferred from company pattern (x)", f"{n_inf} rows. Built from the person's name using the email format proven by a verified/listed email at the same domain (x = pattern, e.g. flast = jsmith@)."],
    ["Best guess (no pattern evidence)", f"{n_guess} rows. No proven format for that domain; uses the most common format in this dataset (first-initial + last name). Alt Email holds the second most likely format."],
    ["Name incomplete", "Only a first name or last initial was available, so no email was inferred."],
    ["", ""],
    ["Phone Type meanings", ""],
    ["Direct (mobile/main)", f"{n_direct} rows. Person-level number from Seamless.AI research."],
    ["Listed (directory/website)", "Number printed next to the person on a directory listing or team page."],
    ["Company main", "Company switchboard number from the directory listing (no person-level number found)."],
    ["", ""],
    ["Priority", "1 = Owner/Principal/President, 2 = VP/GM/C-suite, 3 = Director/Sales Manager, 4 = Outside/Territory Sales, 5 = Inside Sales/Ops/Admin, 6 = Title unknown, 7 = Retired/Former."],
    ["LinkedIn", f"{n_li} rows carry a LinkedIn profile URL (from Seamless.AI records, company team pages, or search results)."],
    ["Data Sources", "seamless = Seamless.AI employee record; websearch = company website/team page or LinkedIn search result; sheet = original directory listing contact."],
    ["", ""],
    ["Caveats", "Seamless.AI sometimes merges unrelated same-named firms under one company record; agents discarded obvious mismatches and flagged doubtful ones in the Companies tab Notes column. Inferred/Best-guess emails are unverified: run them through an email verifier before a large send."],
]
for row in lines:
    ws3.append(row)
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 120
for r in (1, 7, 14):
    ws3.cell(row=r, column=1).font = Font(bold=True)

wb.save(OUT)
print("saved", OUT, os.path.getsize(OUT), "bytes;", len(contacts), "contacts;", len(companies), "companies")
